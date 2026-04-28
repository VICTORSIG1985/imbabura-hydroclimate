"""Extrae las 11 hojas de Tables.xlsx del paquete Wiley a JSON estructurado."""
import json
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).parent.parent
XLSX = ROOT / "scripts" / "_wiley_pkg" / "Tables.xlsx"
OUT = ROOT / "scripts" / "_tables_json"
OUT.mkdir(parents=True, exist_ok=True)

wb = load_workbook(XLSX, data_only=True)
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    rows = []
    for r in ws.iter_rows(values_only=True):
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in r):
            continue
        rows.append([v if v is not None else "" for v in r])
    if len(rows) < 2:
        continue
    caption = rows[0][0] if rows[0] else ""
    headers = [str(h).strip() if h != "" else f"col_{i}" for i, h in enumerate(rows[1])]
    data = [{headers[i]: r[i] if i < len(r) else "" for i in range(len(headers))} for r in rows[2:]]
    fname = sheet_name.replace(" ", "_").replace(".", "") + ".json"
    obj = {
        "sheet": sheet_name,
        "caption": caption.strip() if isinstance(caption, str) else str(caption),
        "n_rows": len(data),
        "n_cols": len(headers),
        "headers": headers,
        "rows": data,
    }
    (OUT / fname).write_text(json.dumps(obj, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    print(f"  {sheet_name:12s} {len(data):4d} rows × {len(headers):3d} cols")
print(f"\nDone: {len(list(OUT.glob('Table_*.json')))} tables in {OUT}")
